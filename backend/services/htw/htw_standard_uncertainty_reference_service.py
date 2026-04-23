from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple
import csv

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from backend.models.htw.htw_standard_uncertainty_reference import (
    HTWStandardUncertaintyReference,
)
from backend.schemas.htw.htw_standard_uncertainty_reference_schemas import (
    HTWStandardUncertaintyReferenceCreate,
    HTWStandardUncertaintyReferenceUpdate,
)

TEMPLATE_COLUMNS: List[str] = [
    "valid_from",
    "valid_upto",
    "torque_nm",
    "applied_torque",
    "indicated_torque",
    "error_value",
    "uncertainty_percent",
]

TEMPLATE_DATE_FORMAT = "YYYY-MM-DD"


# =============================================================================
# CRUD SERVICE LAYER
# =============================================================================
def create_reference(
    db: Session,
    data: HTWStandardUncertaintyReferenceCreate,
) -> HTWStandardUncertaintyReference:
    payload = data.model_dump()
    now = datetime.now(timezone.utc)

    if payload.get("created_at") is None:
        payload["created_at"] = now
    payload["updated_at"] = now

    record = HTWStandardUncertaintyReference(**payload)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def get_references(
    db: Session,
    is_active: Optional[bool] = None,
    torque_nm: Optional[int] = None,
) -> List[HTWStandardUncertaintyReference]:
    query = db.query(HTWStandardUncertaintyReference)

    if is_active is not None:
        query = query.filter(HTWStandardUncertaintyReference.is_active == is_active)

    if torque_nm is not None:
        query = query.filter(HTWStandardUncertaintyReference.torque_nm == torque_nm)

    return query.order_by(HTWStandardUncertaintyReference.created_at.desc()).all()


def get_reference_by_id(
    db: Session,
    reference_id: int,
) -> HTWStandardUncertaintyReference | None:
    return (
        db.query(HTWStandardUncertaintyReference)
        .filter(HTWStandardUncertaintyReference.id == reference_id)
        .first()
    )


def update_reference(
    db: Session,
    reference_id: int,
    data: HTWStandardUncertaintyReferenceUpdate,
) -> HTWStandardUncertaintyReference | None:
    record = get_reference_by_id(db, reference_id)
    if not record:
        return None

    update_data = data.model_dump(exclude_unset=True)

    for key, value in update_data.items():
        setattr(record, key, value)

    record.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(record)
    return record


def delete_reference(
    db: Session,
    reference_id: int,
) -> bool:
    record = get_reference_by_id(db, reference_id)
    if not record:
        return False

    db.delete(record)
    db.commit()
    return True


# =============================================================================
# TEMPLATE / BULK IMPORT HELPERS
# =============================================================================
def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _validate_template_headers(headers: List[Any]) -> None:
    normalized = [_normalize_header(h) for h in headers]
    if normalized != TEMPLATE_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid file template. Expected columns: "
                + ", ".join(TEMPLATE_COLUMNS)
            ),
        )


def _parse_int(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[int]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be an integer")
        return None


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[Decimal]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _parse_date(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[date]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    errors.append(
        f"Row {row_num}: {field_name} must be a valid date in YYYY-MM-DD format"
    )
    return None


def _parse_bool(value: Any, default: bool = True) -> bool:
    if _is_blank(value):
        return default

    raw = str(value).strip().lower()
    if raw in {"true", "1", "yes", "y", "active"}:
        return True
    if raw in {"false", "0", "no", "n", "inactive"}:
        return False
    return default


def _normalize_decimal_for_key(value: Any) -> Decimal:
    """
    Normalize imported and DB decimal values for exact equality checks.
    """
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _build_exact_key(
    valid_from: date,
    valid_upto: date,
    torque_nm: int,
    applied_torque: Any,
    indicated_torque: Any,
    error_value: Any,
    uncertainty_percent: Any,
) -> Tuple[Any, ...]:
    return (
        valid_from,
        valid_upto,
        int(torque_nm),
        _normalize_decimal_for_key(applied_torque),
        _normalize_decimal_for_key(indicated_torque),
        _normalize_decimal_for_key(error_value),
        _normalize_decimal_for_key(uncertainty_percent),
    )


def _fetch_existing_exact_keys(db: Session) -> set[Tuple[Any, ...]]:
    """
    Fetch all currently stored rows and build exact comparison keys based on
    the import/template columns only. This blocks re-importing the same row even
    if the record already exists in the database with a different is_active flag.
    """
    rows = (
        db.query(
            HTWStandardUncertaintyReference.valid_from,
            HTWStandardUncertaintyReference.valid_upto,
            HTWStandardUncertaintyReference.torque_nm,
            HTWStandardUncertaintyReference.applied_torque,
            HTWStandardUncertaintyReference.indicated_torque,
            HTWStandardUncertaintyReference.error_value,
            HTWStandardUncertaintyReference.uncertainty_percent,
        )
        .all()
    )

    existing_keys: set[Tuple[Any, ...]] = set()
    for row in rows:
        existing_keys.add(
            _build_exact_key(
                row.valid_from,
                row.valid_upto,
                row.torque_nm,
                row.applied_torque,
                row.indicated_torque,
                row.error_value,
                row.uncertainty_percent,
            )
        )
    return existing_keys


def _build_rows_from_xlsx(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_template_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(v) for v in row):
            continue

        row_map = {
            TEMPLATE_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(TEMPLATE_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _build_rows_from_csv(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_template_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(v) for v in row.values()):
            continue
        rows.append((row_num, row))
    return rows


def build_template_file(file_format: str = "xlsx") -> tuple[bytes, str, str]:
    """
    Build a downloadable template file for bulk import.
    - Main sheet contains only the mandatory fields.
    - XLSX includes an instructions sheet with date-format guidance.
    """
    normalized = (file_format or "xlsx").strip().lower()

    if normalized == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(TEMPLATE_COLUMNS)

        content = output.getvalue().encode("utf-8-sig")
        return (
            content,
            "text/csv",
            "htw_standard_uncertainty_template.csv",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "HTW Standard Uncertainty"
    ws.append(TEMPLATE_COLUMNS)

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    info = wb.create_sheet(title="Instructions")
    info["A1"] = "Template Notes"
    info["A2"] = "Date format for valid_from and valid_upto"
    info["B2"] = TEMPLATE_DATE_FORMAT
    info["A3"] = "Example"
    info["B3"] = "2026-03-31"
    info["A4"] = "Accepted alternate date formats on import"
    info["B4"] = "DD-MM-YYYY, DD/MM/YYYY"
    info["A6"] = "Keep the main sheet header row exactly as downloaded."
    info["A7"] = "All fields on the main sheet are mandatory."

    output = BytesIO()
    wb.save(output)
    content = output.getvalue()

    return (
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "htw_standard_uncertainty_template.xlsx",
    )


async def import_references_from_upload(
    db: Session,
    file: UploadFile,
) -> Dict[str, Any]:
    """
    Bulk import standard uncertainty references from a predefined template.

    Validation rules:
    - all fields are required
    - dates must be valid
    - torque_nm must be an integer
    - decimal fields must be numeric
    - valid_upto must be >= valid_from
    - exact duplicate rows are blocked:
        * duplicate inside the same uploaded file
        * duplicate already existing in the database
    """
    filename = (file.filename or "").lower().strip()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx and .csv files are allowed",
        )

    content = await file.read()

    if filename.endswith(".xlsx"):
        rows = _build_rows_from_xlsx(content)
    else:
        rows = _build_rows_from_csv(content)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No data rows found in the uploaded file",
        )

    errors: List[str] = []
    objects: List[HTWStandardUncertaintyReference] = []
    seen_keys: set[Tuple[Any, ...]] = set()
    existing_keys = _fetch_existing_exact_keys(db)
    now = datetime.now(timezone.utc)

    for row_num, row in rows:
        valid_from = _parse_date(row.get("valid_from"), "valid_from", row_num, errors)
        valid_upto = _parse_date(row.get("valid_upto"), "valid_upto", row_num, errors)
        torque_nm = _parse_int(row.get("torque_nm"), "torque_nm", row_num, errors)
        applied_torque = _parse_decimal(
            row.get("applied_torque"), "applied_torque", row_num, errors
        )
        indicated_torque = _parse_decimal(
            row.get("indicated_torque"), "indicated_torque", row_num, errors
        )
        error_value = _parse_decimal(
            row.get("error_value"), "error_value", row_num, errors
        )
        uncertainty_percent = _parse_decimal(
            row.get("uncertainty_percent"), "uncertainty_percent", row_num, errors
        )

        if valid_from and valid_upto and valid_upto < valid_from:
            errors.append(
                f"Row {row_num}: valid_upto must be greater than or equal to valid_from"
            )

        # Only build the duplicate-check key when all required values are valid.
        row_key = None
        if (
            valid_from is not None
            and valid_upto is not None
            and torque_nm is not None
            and applied_torque is not None
            and indicated_torque is not None
            and error_value is not None
            and uncertainty_percent is not None
        ):
            row_key = _build_exact_key(
                valid_from,
                valid_upto,
                torque_nm,
                applied_torque,
                indicated_torque,
                error_value,
                uncertainty_percent,
            )

            if row_key in seen_keys:
                errors.append(
                    f"Row {row_num}: duplicate row found inside the uploaded file"
                )
            elif row_key in existing_keys:
                errors.append(
                    f"Row {row_num}: duplicate row already exists in the database"
                )
            else:
                seen_keys.add(row_key)

        objects.append(
            HTWStandardUncertaintyReference(
                valid_from=valid_from,
                valid_upto=valid_upto,
                torque_nm=torque_nm,
                applied_torque=applied_torque,
                indicated_torque=indicated_torque,
                error_value=error_value,
                uncertainty_percent=uncertainty_percent,
                is_active=True,
                created_at=now,
                updated_at=now,
            )
        )

    if errors:
        raise HTTPException(
            status_code=400,
            detail="Import failed:\n" + "\n".join(errors[:100]),
        )

    try:
        db.add_all(objects)
        db.commit()
        return {
            "message": "Import successful",
            "inserted_rows": len(objects),
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Database error: {str(exc)}",
        )
