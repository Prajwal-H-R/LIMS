from __future__ import annotations

import csv
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from backend.models.htw.htw_cmc_reference import HTWCMCReference
from backend.schemas.htw.htw_cmc_reference_schema import (
    HTWCMCReferenceCreate,
    HTWCMCReferenceUpdate,
)

TEMPLATE_COLUMNS: List[str] = [
    "lower_measure_range",
    "higher_measure_range",
    "cmc_percent",
]


# =============================================================================
# CRUD SERVICE LAYER
# =============================================================================
def create_cmc_reference(db: Session, payload: HTWCMCReferenceCreate) -> HTWCMCReference:
    data = payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()
    data.setdefault("is_active", True)

    now = datetime.now(timezone.utc)
    if "created_at" in HTWCMCReference.__table__.columns:
        data.setdefault("created_at", now)
    if "updated_at" in HTWCMCReference.__table__.columns:
        data["updated_at"] = now

    obj = HTWCMCReference(**data)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def get_all_cmc_references(db: Session) -> List[HTWCMCReference]:
    return (
        db.query(HTWCMCReference)
        .order_by(HTWCMCReference.lower_measure_range.asc())
        .all()
    )


def get_cmc_reference_by_id(db: Session, record_id: int) -> HTWCMCReference | None:
    return (
        db.query(HTWCMCReference)
        .filter(HTWCMCReference.id == record_id)
        .first()
    )


def get_applicable_cmc(db: Session, applied_value: float):
    return (
        db.query(HTWCMCReference)
        .filter(
            HTWCMCReference.is_active.is_(True),
            HTWCMCReference.lower_measure_range <= applied_value,
            HTWCMCReference.higher_measure_range >= applied_value,
        )
        .order_by(HTWCMCReference.higher_measure_range.asc())
        .first()
    )


def update_cmc_reference(
    db: Session,
    record_id: int,
    payload: HTWCMCReferenceUpdate,
):
    obj = get_cmc_reference_by_id(db, record_id)
    if not obj:
        return None

    update_data = payload.model_dump(exclude_unset=True) if hasattr(payload, "model_dump") else payload.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(obj, key, value)

    if hasattr(obj, "updated_at"):
        obj.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(obj)
    return obj


def deactivate_cmc_reference(db: Session, record_id: int) -> bool:
    obj = get_cmc_reference_by_id(db, record_id)
    if not obj:
        return False

    obj.is_active = False
    if hasattr(obj, "updated_at"):
        obj.updated_at = datetime.now(timezone.utc)

    db.commit()
    return True


# =============================================================================
# TEMPLATE / BULK IMPORT HELPERS
# =============================================================================
def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _validate_template_headers(headers: List[Any]) -> None:
    normalized = [_normalize_header(h) for h in headers]
    if normalized != TEMPLATE_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail="Invalid file template. Expected columns: " + ", ".join(TEMPLATE_COLUMNS),
        )


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[Decimal]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _build_rows_from_xlsx(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_template_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(v) for v in row):
            continue

        row_map = {
            TEMPLATE_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(TEMPLATE_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _build_rows_from_csv(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_template_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(v) for v in row.values()):
            continue
        rows.append((row_num, row))
    return rows


def build_template_file(file_format: str = "xlsx") -> tuple[bytes, str, str]:
    """
    Build a downloadable template file for bulk import.
    XLSX includes a small instructions sheet.
    """
    normalized = (file_format or "xlsx").strip().lower()

    if normalized == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(TEMPLATE_COLUMNS)
        return (
            output.getvalue().encode("utf-8-sig"),
            "text/csv",
            "htw_cmc_template.csv",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "CMC Scope"
    ws.append(TEMPLATE_COLUMNS)

    info = wb.create_sheet(title="Instructions")
    info["A1"] = "Template Notes"
    info["A2"] = "All fields in the main sheet are mandatory."
    info["A3"] = "Use numeric values for all columns."
    info["A4"] = "Example lower_measure_range"
    info["B4"] = "200"
    info["A5"] = "Example higher_measure_range"
    info["B5"] = "1500"
    info["A6"] = "Example cmc_percent"
    info["B6"] = "0.58"
    info["A8"] = "Keep the main sheet header row exactly as downloaded."

    output = BytesIO()
    wb.save(output)
    return (
        output.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "htw_cmc_template.xlsx",
    )


def _db_key(value: Any) -> Decimal:
    return Decimal(str(value).strip())


def import_cmc_references_from_upload(
    db: Session,
    file: UploadFile,
) -> Dict[str, Any]:
    """
    Bulk import CMC references from a predefined template.
    Rejects duplicate rows both within the file and against the database.
    """
    filename = (file.filename or "").lower().strip()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are allowed")

    content = file.file.read() if getattr(file, "file", None) else None
    if content is None:
        raise HTTPException(status_code=400, detail="Unable to read uploaded file")

    if filename.endswith(".xlsx"):
        rows = _build_rows_from_xlsx(content)
    else:
        rows = _build_rows_from_csv(content)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")

    existing_keys = {
        (
            _db_key(obj.lower_measure_range),
            _db_key(obj.higher_measure_range),
            _db_key(obj.cmc_percent),
        )
        for obj in db.query(HTWCMCReference).all()
    }

    errors: List[str] = []
    seen_in_file: set[tuple[Decimal, Decimal, Decimal]] = set()
    objects: List[HTWCMCReference] = []

    for row_num, row in rows:
        lower = _parse_decimal(row.get("lower_measure_range"), "lower_measure_range", row_num, errors)
        higher = _parse_decimal(row.get("higher_measure_range"), "higher_measure_range", row_num, errors)
        cmc = _parse_decimal(row.get("cmc_percent"), "cmc_percent", row_num, errors)

        if lower is not None and higher is not None and lower >= higher:
            errors.append(f"Row {row_num}: lower_measure_range must be less than higher_measure_range")

        if lower is not None and higher is not None and cmc is not None:
            row_key = (lower, higher, cmc)
            if row_key in seen_in_file:
                errors.append(
                    f"Row {row_num}: duplicate row found inside the uploaded file "
                    "(same lower_measure_range, higher_measure_range, cmc_percent)"
                )
            if row_key in existing_keys:
                errors.append(
                    f"Row {row_num}: duplicate row already exists in the database "
                    "(same lower_measure_range, higher_measure_range, cmc_percent)"
                )
            seen_in_file.add(row_key)

        objects.append(
            HTWCMCReference(
                lower_measure_range=lower,
                higher_measure_range=higher,
                cmc_percent=cmc,
                is_active=True,
            )
        )

    if errors:
        raise HTTPException(
            status_code=400,
            detail="Import failed:\n" + "\n".join(errors[:100]),
        )

    try:
        db.add_all(objects)
        db.commit()
        return {
            "message": "Import successful",
            "inserted_rows": len(objects),
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(exc)}")
