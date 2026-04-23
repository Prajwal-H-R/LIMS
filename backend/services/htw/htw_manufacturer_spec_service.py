from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from backend import models
from backend.schemas.htw.htw_manufacturer_spec_schemas import (
    HTWManufacturerSpecCreate,
    HTWManufacturerSpecUpdate,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


EXPECTED_COLUMNS = [
    "make",
    "model",
    "range_min",
    "range_max",
    "torque_20",
    "torque_40",
    "torque_60",
    "torque_80",
    "torque_100",
    "torque_unit",
    "pressure_20",
    "pressure_40",
    "pressure_60",
    "pressure_80",
    "pressure_100",
    "pressure_unit",
]


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _clean_text(value: Any) -> Optional[str]:
    if _is_blank(value):
        return None
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
    required: bool,
) -> Optional[Decimal]:
    if _is_blank(value):
        if required:
            errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _validate_headers(headers: List[Any]) -> None:
    normalized_headers = [_normalize_header(h) for h in headers]
    if normalized_headers != EXPECTED_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid file template. Headers must match exactly: "
                + ", ".join(EXPECTED_COLUMNS)
            ),
        )


def _existing_make_model_pairs(db: Session) -> set[Tuple[str, str]]:
    existing_rows = db.query(
        models.HTWManufacturerSpec.make,
        models.HTWManufacturerSpec.model,
    ).all()

    pairs: set[Tuple[str, str]] = set()
    for make, model in existing_rows:
        if make is None or model is None:
            continue
        pairs.add((str(make).strip().lower(), str(model).strip().lower()))
    return pairs


def _read_xlsx_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    if load_workbook is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server",
        )

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(cell) for cell in row):
            continue

        row_map = {
            EXPECTED_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(EXPECTED_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _read_csv_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(cell) for cell in row.values()):
            continue

        row_map = {k: v for k, v in row.items()}
        rows.append((row_num, row_map))

    return rows


def get_htw_manufacturer_specs(
    db: Session,
    skip: int = 0,
    limit: Optional[int] = None,
    is_active: Optional[bool] = None,
):
    query = db.query(models.HTWManufacturerSpec)

    if is_active is not None:
        query = query.filter(models.HTWManufacturerSpec.is_active == is_active)

    query = query.order_by(models.HTWManufacturerSpec.created_at.desc())
    return query.offset(skip).limit(limit).all()


def get_htw_manufacturer_spec(db: Session, spec_id: int):
    spec = db.query(models.HTWManufacturerSpec).filter(
        models.HTWManufacturerSpec.id == spec_id
    ).first()

    if not spec:
        raise HTTPException(
            status_code=404,
            detail=f"HTW Manufacturer Spec with ID {spec_id} not found",
        )

    return spec


def create_htw_manufacturer_spec(
    db: Session,
    spec_data: HTWManufacturerSpecCreate,
):
    try:
        new_spec = models.HTWManufacturerSpec(**spec_data.model_dump())
        db.add(new_spec)
        db.commit()
        db.refresh(new_spec)
        return new_spec
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"An internal server error occurred: {str(e)}",
        )


def update_htw_manufacturer_spec(
    db: Session,
    spec_id: int,
    spec_data: HTWManufacturerSpecUpdate,
):
    try:
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found",
            )

        update_data = spec_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(spec, field, value)

        from datetime import datetime
        spec.updated_at = datetime.now()

        db.commit()
        db.refresh(spec)
        return spec
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"An internal server error occurred: {str(e)}",
        )


def update_htw_manufacturer_spec_status(
    db: Session,
    spec_id: int,
    is_active: bool,
):
    try:
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found",
            )

        from datetime import datetime
        spec.is_active = is_active
        spec.updated_at = datetime.now()

        db.commit()
        db.refresh(spec)
        return spec
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"An internal server error occurred: {str(e)}",
        )


def delete_htw_manufacturer_spec(db: Session, spec_id: int):
    try:
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found",
            )

        db.delete(spec)
        db.commit()
        return None
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"An internal server error occurred: {str(e)}",
        )


def download_htw_manufacturer_spec_template(file_format: str = "xlsx"):
    if file_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPECTED_COLUMNS)

        csv_data = output.getvalue().encode("utf-8-sig")
        headers = {
            "Content-Disposition": 'attachment; filename="htw_manufacturer_spec_template.csv"'
        }
        return Response(content=csv_data, media_type="text/csv", headers=headers)

    if Workbook is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Manufacturer Specs"
    sheet.append(EXPECTED_COLUMNS)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="htw_manufacturer_spec_template.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


async def import_htw_manufacturer_specs(
    db: Session,
    filename: str,
    file_bytes: bytes,
):
    filename = (filename or "").lower()
    if not filename.endswith(".xlsx") and not filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are allowed")

    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if filename.endswith(".xlsx"):
        rows = _read_xlsx_rows(file_bytes)
    else:
        rows = _read_csv_rows(file_bytes)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")

    errors: List[str] = []
    existing_pairs = _existing_make_model_pairs(db)
    seen_in_file: set[Tuple[str, str]] = set()
    objects_to_insert = []

    for row_num, row in rows:
        make = _clean_text(row.get("make"))
        model = _clean_text(row.get("model"))
        torque_unit = _clean_text(row.get("torque_unit"))
        pressure_unit = _clean_text(row.get("pressure_unit"))

        if not make:
            errors.append(f"Row {row_num}: make is required")
        if not model:
            errors.append(f"Row {row_num}: model is required")
        if not torque_unit:
            errors.append(f"Row {row_num}: torque_unit is required")
        if not pressure_unit:
            errors.append(f"Row {row_num}: pressure_unit is required")

        key = ((make or "").strip().lower(), (model or "").strip().lower())
        if make and model:
            if key in seen_in_file:
                errors.append(f"Row {row_num}: duplicate make + model inside the uploaded file")
            if key in existing_pairs:
                errors.append(f"Row {row_num}: make + model already exists in the database")
            seen_in_file.add(key)

        parsed_data = {
            "make": make,
            "model": model,
            "range_min": _parse_decimal(row.get("range_min"), "range_min", row_num, errors, True),
            "range_max": _parse_decimal(row.get("range_max"), "range_max", row_num, errors, True),
            "torque_20": _parse_decimal(row.get("torque_20"), "torque_20", row_num, errors, True),
            "torque_40": _parse_decimal(row.get("torque_40"), "torque_40", row_num, errors, False),
            "torque_60": _parse_decimal(row.get("torque_60"), "torque_60", row_num, errors, True),
            "torque_80": _parse_decimal(row.get("torque_80"), "torque_80", row_num, errors, False),
            "torque_100": _parse_decimal(row.get("torque_100"), "torque_100", row_num, errors, True),
            "torque_unit": torque_unit,
            "pressure_20": _parse_decimal(row.get("pressure_20"), "pressure_20", row_num, errors, True),
            "pressure_40": _parse_decimal(row.get("pressure_40"), "pressure_40", row_num, errors, False),
            "pressure_60": _parse_decimal(row.get("pressure_60"), "pressure_60", row_num, errors, True),
            "pressure_80": _parse_decimal(row.get("pressure_80"), "pressure_80", row_num, errors, False),
            "pressure_100": _parse_decimal(row.get("pressure_100"), "pressure_100", row_num, errors, True),
            "pressure_unit": pressure_unit,
            "is_active": True,
        }

        range_min = parsed_data["range_min"]
        range_max = parsed_data["range_max"]
        if range_min is not None and range_max is not None and range_max < range_min:
            errors.append(f"Row {row_num}: range_max must be greater than or equal to range_min")

        objects_to_insert.append(models.HTWManufacturerSpec(**parsed_data))

    if errors:
        raise HTTPException(
            status_code=400,
            detail="Import failed:\n" + "\n".join(errors[:100]),
        )

    try:
        db.add_all(objects_to_insert)
        db.commit()
        return {
            "message": "Import successful",
            "inserted_rows": len(objects_to_insert),
        }
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"An internal server error occurred: {str(e)}",
        )