from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from backend import models
from backend.schemas.htw.htw_t_distribution_schema import (
    HTWTDistributionCreate,
    HTWTDistributionUpdate,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


EXPECTED_COLUMNS = [
    "degrees_of_freedom",
    "confidence_level",
    "alpha",
    "t_value",
]


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _validate_headers(headers: List[Any]) -> None:
    normalized = [_normalize_header(h) for h in headers]
    if normalized != EXPECTED_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail="Invalid file template. Headers must match exactly: " + ", ".join(EXPECTED_COLUMNS),
        )


def _parse_int(value: Any, field_name: str, row_num: int, errors: List[str]) -> Optional[int]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        errors.append(f"Row {row_num}: {field_name} must be an integer")
        return None


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[Decimal]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None
    try:
        return Decimal(str(value).strip())
    except (TypeError, ValueError, InvalidOperation):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _read_xlsx_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(cell) for cell in row):
            continue

        row_map = {
            EXPECTED_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(EXPECTED_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _read_csv_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(cell) for cell in row.values()):
            continue

        row_map = {k: v for k, v in row.items()}
        rows.append((row_num, row_map))

    return rows


def _existing_pairs(db: Session) -> set[Tuple[int, Decimal]]:
    rows = db.query(
        models.HTWTDistribution.degrees_of_freedom,
        models.HTWTDistribution.confidence_level,
    ).all()

    pairs: set[Tuple[int, Decimal]] = set()
    for degrees_of_freedom, confidence_level in rows:
        if degrees_of_freedom is None or confidence_level is None:
            continue
        try:
            pairs.add((int(degrees_of_freedom), Decimal(str(confidence_level))))
        except Exception:
            continue
    return pairs


def download_t_distribution_template(file_format: str = "xlsx"):
    file_format = (file_format or "xlsx").lower().strip()
    if file_format not in {"xlsx", "csv"}:
        raise HTTPException(status_code=400, detail="file_format must be either xlsx or csv")

    if file_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPECTED_COLUMNS)

        csv_data = output.getvalue().encode("utf-8-sig")
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="htw_t_distribution_template.csv"'},
        )

    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T Distribution"
    sheet.append(EXPECTED_COLUMNS)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="htw_t_distribution_template.xlsx"'},
    )


async def import_t_distribution(
    db: Session,
    filename: str,
    file_bytes: bytes,
):
    filename = (filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are allowed")

    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if filename.endswith(".xlsx"):
        rows = _read_xlsx_rows(file_bytes)
    else:
        rows = _read_csv_rows(file_bytes)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")

    errors: List[str] = []
    existing_pairs = _existing_pairs(db)
    seen_in_file: set[Tuple[int, Decimal]] = set()
    objects_to_insert = []

    for row_num, row in rows:
        degrees_of_freedom = _parse_int(row.get("degrees_of_freedom"), "degrees_of_freedom", row_num, errors)
        confidence_level = _parse_decimal(row.get("confidence_level"), "confidence_level", row_num, errors)
        alpha = _parse_decimal(row.get("alpha"), "alpha", row_num, errors)
        t_value = _parse_decimal(row.get("t_value"), "t_value", row_num, errors)

        if degrees_of_freedom is not None and degrees_of_freedom <= 0:
            errors.append(f"Row {row_num}: degrees_of_freedom must be greater than 0")

        if confidence_level is not None and (confidence_level < 0 or confidence_level > 100):
            errors.append(f"Row {row_num}: confidence_level must be between 0 and 100")

        if alpha is not None and alpha < 0:
            errors.append(f"Row {row_num}: alpha must be greater than or equal to 0")

        if t_value is not None and t_value < 0:
            errors.append(f"Row {row_num}: t_value must be greater than or equal to 0")

        if degrees_of_freedom is not None and confidence_level is not None:
            key = (degrees_of_freedom, confidence_level)
            if key in seen_in_file:
                errors.append(f"Row {row_num}: duplicate degrees_of_freedom + confidence_level inside the uploaded file")
            if key in existing_pairs:
                errors.append(f"Row {row_num}: degrees_of_freedom + confidence_level already exists in the database")
            seen_in_file.add(key)

        if None in (degrees_of_freedom, confidence_level, alpha, t_value):
            continue

        objects_to_insert.append(
            models.HTWTDistribution(
                degrees_of_freedom=degrees_of_freedom,
                confidence_level=confidence_level,
                alpha=alpha,
                t_value=t_value,
                is_active=True,
            )
        )

    if errors:
        raise HTTPException(status_code=400, detail="Import failed:\n" + "\n".join(errors[:100]))

    try:
        db.add_all(objects_to_insert)
        db.commit()
        return {"message": "Import successful", "inserted_rows": len(objects_to_insert)}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


def create_t_distribution(
    db: Session,
    payload: HTWTDistributionCreate,
):
    try:
        record = models.HTWTDistribution(**payload.dict())
        db.add(record)
        db.commit()
        db.refresh(record)
        return record
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="T-distribution entry already exists for given DF and confidence level",
        )


def get_t_distribution_by_id(
    db: Session,
    record_id: int,
):
    record = (
        db.query(models.HTWTDistribution)
        .filter(models.HTWTDistribution.id == record_id)
        .first()
    )

    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="T-distribution record not found",
        )

    return record


def list_t_distributions(
    db: Session,
    active_only: bool = True,
):
    query = db.query(models.HTWTDistribution)

    if active_only:
        query = query.filter(models.HTWTDistribution.is_active.is_(True))

    return query.order_by(
        models.HTWTDistribution.degrees_of_freedom.asc(),
        models.HTWTDistribution.confidence_level.asc(),
    ).all()


def update_t_distribution(
    db: Session,
    record_id: int,
    payload: HTWTDistributionUpdate,
):
    record = get_t_distribution_by_id(db, record_id)

    for field, value in payload.dict(exclude_unset=True).items():
        setattr(record, field, value)

    try:
        db.commit()
        db.refresh(record)
        return record
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Duplicate T-distribution entry after update",
        )


def soft_delete_t_distribution(
    db: Session,
    record_id: int,
):
    record = get_t_distribution_by_id(db, record_id)
    record.is_active = False
    db.commit()