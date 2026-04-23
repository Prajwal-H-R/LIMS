from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple
import csv

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from backend import models
from backend.schemas.htw.htw_pressure_gauge_resolution import (
    HTWPressureGaugeResolutionCreate,
)

# ✅ Removed "is_active" from template columns
TEMPLATE_COLUMNS: List[str] = [
    "pressure",
    "unit",
    "valid_upto",
]

TEMPLATE_DATE_FORMAT = "YYYY-MM-DD"


# =============================================================================
# CRUD SERVICE LAYER
# =============================================================================
def create_resolution(
    db: Session,
    data: HTWPressureGaugeResolutionCreate,
) -> models.HTWPressureGaugeResolution:
    payload = data.model_dump()
    now = datetime.now(timezone.utc)

    if payload.get("created_at") is None:
        payload["created_at"] = now

    # ✅ Only add updated_at if the model actually defines it
    if hasattr(models.HTWPressureGaugeResolution, "updated_at"):
        payload["updated_at"] = now

    # ✅ Ensure is_active is always True
    payload["is_active"] = True

    record = models.HTWPressureGaugeResolution(**payload)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def get_resolutions(
    db: Session,
    unit: Optional[str] = None,
) -> List[models.HTWPressureGaugeResolution]:
    query = db.query(models.HTWPressureGaugeResolution)

    if unit:
        query = query.filter(
            models.HTWPressureGaugeResolution.unit.isnot(None),
            models.HTWPressureGaugeResolution.unit == unit.strip(),
        )

    return query.order_by(models.HTWPressureGaugeResolution.pressure.asc()).all()


def get_unique_units(db: Session) -> List[str]:
    units = (
        db.query(models.HTWPressureGaugeResolution.unit)
        .filter(models.HTWPressureGaugeResolution.unit.isnot(None))
        .distinct()
        .order_by(models.HTWPressureGaugeResolution.unit.asc())
        .all()
    )
    return [unit[0] for unit in units if unit and unit[0]]


def get_resolution_by_id(
    db: Session,
    resolution_id: int,
) -> models.HTWPressureGaugeResolution | None:
    return (
        db.query(models.HTWPressureGaugeResolution)
        .filter(models.HTWPressureGaugeResolution.id == resolution_id)
        .first()
    )


def update_resolution(
    db: Session,
    resolution_id: int,
    data: HTWPressureGaugeResolutionCreate,
) -> models.HTWPressureGaugeResolution | None:
    record = get_resolution_by_id(db, resolution_id)
    if not record:
        return None

    update_data = data.model_dump(exclude_unset=True)

    for key, value in update_data.items():
        setattr(record, key, value)

    if hasattr(record, "updated_at"):
        record.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(record)
    return record


def update_resolution_status(
    db: Session,
    resolution_id: int,
    is_active: bool,
) -> models.HTWPressureGaugeResolution | None:
    record = get_resolution_by_id(db, resolution_id)
    if not record:
        return None

    record.is_active = is_active
    if hasattr(record, "updated_at"):
        record.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(record)
    return record


def delete_resolution(
    db: Session,
    resolution_id: int,
) -> bool:
    record = get_resolution_by_id(db, resolution_id)
    if not record:
        return False

    db.delete(record)
    db.commit()
    return True


# =============================================================================
# TEMPLATE / BULK IMPORT HELPERS
# =============================================================================
def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _validate_template_headers(headers: List[Any]) -> None:
    normalized = [_normalize_header(h) for h in headers]
    if normalized == TEMPLATE_COLUMNS:
        return
    raise HTTPException(
        status_code=400,
        detail=(
            "Invalid file template. Expected columns: "
            + ", ".join(TEMPLATE_COLUMNS)
        ),
    )


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[Decimal]:
    if _is_blank(value):
        errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _parse_date(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
) -> Optional[date]:
    if _is_blank(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    errors.append(
        f"Row {row_num}: {field_name} must be a valid date in YYYY-MM-DD format"
    )
    return None


def _build_rows_from_xlsx(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_template_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(v) for v in row):
            continue

        row_map = {
            TEMPLATE_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(TEMPLATE_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _build_rows_from_csv(content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_template_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(v) for v in row.values()):
            continue
        rows.append((row_num, row))
    return rows


def build_template_file(file_format: str = "xlsx") -> tuple[bytes, str, str]:
    normalized = (file_format or "xlsx").strip().lower()

    if normalized == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(TEMPLATE_COLUMNS)

        content = output.getvalue().encode("utf-8-sig")
        return (
            content,
            "text/csv",
            "htw_pressure_gauge_resolution_template.csv",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Pressure Gauge Resolutions"
    ws.append(TEMPLATE_COLUMNS)

    info = wb.create_sheet(title="Instructions")
    info["A1"] = "Template Notes"
    info["A2"] = "Date format for valid_upto"
    info["B2"] = TEMPLATE_DATE_FORMAT
    info["A3"] = "Example"
    info["B3"] = "2026-03-31"
    info["A4"] = "Accepted alternate date formats on import"
    info["B4"] = "DD-MM-YYYY, DD/MM/YYYY"
    info["A5"] = "valid_upto can be left blank for indefinite validity"
    info["A6"] = "Keep the main sheet header row exactly as downloaded."

    output = BytesIO()
    wb.save(output)
    content = output.getvalue()

    return (
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "htw_pressure_gauge_resolution_template.xlsx",
    )


def _existing_import_keys(db: Session) -> set[tuple[Decimal, str, Optional[date]]]:
    rows = db.query(
        models.HTWPressureGaugeResolution.pressure,
        models.HTWPressureGaugeResolution.unit,
        models.HTWPressureGaugeResolution.valid_upto,
    ).all()

    keys: set[tuple[Decimal, str, Optional[date]]] = set()
    for pressure, unit, valid_upto in rows:
        if pressure is None or unit is None:
            continue
        pressure_key = Decimal(str(pressure))
        unit_key = str(unit).strip().lower()
        valid_upto_key = valid_upto
        if isinstance(valid_upto_key, datetime):
            valid_upto_key = valid_upto_key.date()
        keys.add((pressure_key, unit_key, valid_upto_key))
    return keys


async def import_resolutions_from_upload(
    db: Session,
    file: UploadFile,
) -> Dict[str, Any]:
    filename = (file.filename or "").lower().strip()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx and .csv files are allowed",
        )

    content = await file.read()

    if filename.endswith(".xlsx"):
        rows = _build_rows_from_xlsx(content)
    else:
        rows = _build_rows_from_csv(content)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No data rows found in the uploaded file",
        )

    existing_keys = _existing_import_keys(db)
    seen_in_file: set[tuple[Decimal, str, Optional[date]]] = set()
    parsed_rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    now = datetime.now(timezone.utc)

    for row_num, row in rows:
        pressure = _parse_decimal(row.get("pressure"), "pressure", row_num, errors)
        unit = row.get("unit")
        valid_upto = _parse_date(row.get("valid_upto"), "valid_upto", row_num, errors)
        
        if _is_blank(unit):
            errors.append(f"Row {row_num}: unit is required")
            unit = None

        if pressure is not None and unit is not None:
            key = (pressure, str(unit).strip().lower(), valid_upto)
            if key in seen_in_file:
                errors.append(
                    f"Row {row_num}: duplicate row inside the uploaded file "
                    f"(pressure, unit, valid_upto already repeated)"
                )
            if key in existing_keys:
                errors.append(
                    f"Row {row_num}: duplicate row already exists in the database "
                    f"(pressure, unit, valid_upto)"
                )
            seen_in_file.add(key)

        # ✅ Build row dict without updated_at to match model definition
        row_data = {
            "pressure": pressure,
            "unit": str(unit).strip() if unit is not None else None,
            "valid_upto": valid_upto,
            "is_active": True,
            "created_at": now,
        }
        if hasattr(models.HTWPressureGaugeResolution, "updated_at"):
            row_data["updated_at"] = now
            
        parsed_rows.append(row_data)

    if errors:
        raise HTTPException(
            status_code=400,
            detail="Import failed:\n" + "\n".join(errors[:100]),
        )

    try:
        db.add_all([models.HTWPressureGaugeResolution(**item) for item in parsed_rows])
        db.commit()
        return {
            "message": "Import successful",
            "inserted_rows": len(parsed_rows),
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Database error: {str(exc)}",
        )