from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from typing import List, Optional, Any, Dict, Tuple
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import csv

from backend.schemas.htw.htw_manufacturer_spec_schemas import (
    HTWManufacturerSpecCreate,
    HTWManufacturerSpecUpdate,
    HTWManufacturerSpecResponse,
)
from backend import models
from backend.db import get_db
from backend.auth import get_current_user
from backend.schemas.user_schemas import UserResponse

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


router = APIRouter(
    prefix="/htw-manufacturer-specs",
    tags=["HTW Manufacturer Specifications"]
)


# =====================================================================
# GET: All HTW Manufacturer Specs (List View)
# =====================================================================
@router.get("/", response_model=List[HTWManufacturerSpecResponse])
def get_htw_manufacturer_specs(
    db: Session = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: Optional[int] = Query(None, ge=1),  # no upper bound
    is_active: Optional[bool] = Query(None)
):
    """
    Retrieves a list of HTW Manufacturer Specifications.
    Only returns specs for Hydraulic Torque Wrench equipment type.
    """
    try:
        query = db.query(models.HTWManufacturerSpec)

        # Filter by active status if provided
        if is_active is not None:
            query = query.filter(models.HTWManufacturerSpec.is_active == is_active)

        # Order by created_at descending (newest first)
        query = query.order_by(models.HTWManufacturerSpec.created_at.desc())

        # Apply pagination
        specs = query.offset(skip).limit(limit).all()

        return specs
    except SQLAlchemyError as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# IMPORT HELPERS
# =====================================================================
EXPECTED_COLUMNS = [
    "make",
    "model",
    "range_min",
    "range_max",
    "torque_20",
    "torque_40",
    "torque_60",
    "torque_80",
    "torque_100",
    "torque_unit",
    "pressure_20",
    "pressure_40",
    "pressure_60",
    "pressure_80",
    "pressure_100",
    "pressure_unit",
    "is_active",
]


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _clean_text(value: Any) -> Optional[str]:
    if _is_blank(value):
        return None
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _parse_decimal(
    value: Any,
    field_name: str,
    row_num: int,
    errors: List[str],
    required: bool,
) -> Optional[Decimal]:
    if _is_blank(value):
        if required:
            errors.append(f"Row {row_num}: {field_name} is required")
        return None

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"Row {row_num}: {field_name} must be numeric")
        return None


def _parse_bool(value: Any, row_num: int, errors: List[str]) -> bool:
    if _is_blank(value):
        return True

    raw = str(value).strip().lower()
    if raw in {"true", "1", "yes", "y", "active"}:
        return True
    if raw in {"false", "0", "no", "n", "inactive"}:
        return False

    errors.append(f"Row {row_num}: is_active must be TRUE/FALSE, YES/NO, 1/0")
    return True


def _validate_headers(headers: List[Any]) -> None:
    normalized_headers = [_normalize_header(h) for h in headers]
    if normalized_headers != EXPECTED_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid file template. "
                "Headers must match exactly: "
                + ", ".join(EXPECTED_COLUMNS)
            ),
        )


def _existing_make_model_pairs(db: Session) -> set[Tuple[str, str]]:
    existing_rows = db.query(
        models.HTWManufacturerSpec.make,
        models.HTWManufacturerSpec.model,
    ).all()

    pairs: set[Tuple[str, str]] = set()
    for make, model in existing_rows:
        if make is None or model is None:
            continue
        pairs.add((str(make).strip().lower(), str(model).strip().lower()))
    return pairs


def _read_xlsx_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    if load_workbook is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server",
        )

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty")

    _validate_headers(list(header_row))

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(_is_blank(cell) for cell in row):
            continue

        row_map = {
            EXPECTED_COLUMNS[i]: (row[i] if i < len(row) else None)
            for i in range(len(EXPECTED_COLUMNS))
        }
        rows.append((row_num, row_map))

    return rows


def _read_csv_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV file is empty")

    _validate_headers(reader.fieldnames)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, row in enumerate(reader, start=2):
        if row is None or all(_is_blank(cell) for cell in row.values()):
            continue

        row_map = {k: v for k, v in row.items()}
        rows.append((row_num, row_map))

    return rows


# =====================================================================
# NEW: Download fixed template for bulk import
# =====================================================================
@router.get("/template")
def download_htw_manufacturer_spec_template(
    file_format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Downloads a fixed manufacturer specification template in XLSX or CSV format.
    """
    if file_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPECTED_COLUMNS)

        csv_data = output.getvalue().encode("utf-8-sig")
        headers = {
            "Content-Disposition": 'attachment; filename="htw_manufacturer_spec_template.csv"'
        }
        return Response(content=csv_data, media_type="text/csv", headers=headers)

    if Workbook is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Manufacturer Specs"
    sheet.append(EXPECTED_COLUMNS)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="htw_manufacturer_spec_template.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# =====================================================================
# NEW: Bulk import manufacturer specs from XLSX / CSV
# =====================================================================
@router.post("/import")
async def import_htw_manufacturer_specs(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Bulk imports HTW manufacturer specs from a fixed template.
    This is all-or-nothing: if any row has an error, the import is rejected.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx") and not filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are allowed")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if filename.endswith(".xlsx"):
        rows = _read_xlsx_rows(file_bytes)
    else:
        rows = _read_csv_rows(file_bytes)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")

    errors: List[str] = []
    existing_pairs = _existing_make_model_pairs(db)
    seen_in_file: set[Tuple[str, str]] = set()
    objects_to_insert = []

    for row_num, row in rows:
        make = _clean_text(row.get("make"))
        model = _clean_text(row.get("model"))
        torque_unit = _clean_text(row.get("torque_unit"))
        pressure_unit = _clean_text(row.get("pressure_unit"))

        if not make:
            errors.append(f"Row {row_num}: make is required")
        if not model:
            errors.append(f"Row {row_num}: model is required")
        if not torque_unit:
            errors.append(f"Row {row_num}: torque_unit is required")
        if not pressure_unit:
            errors.append(f"Row {row_num}: pressure_unit is required")

        key = ((make or "").strip().lower(), (model or "").strip().lower())
        if make and model:
            if key in seen_in_file:
                errors.append(f"Row {row_num}: duplicate make + model inside the uploaded file")
            if key in existing_pairs:
                errors.append(f"Row {row_num}: make + model already exists in the database")
            seen_in_file.add(key)

        parsed_data = {
            "make": make,
            "model": model,
            "range_min": _parse_decimal(row.get("range_min"), "range_min", row_num, errors, True),
            "range_max": _parse_decimal(row.get("range_max"), "range_max", row_num, errors, True),
            "torque_20": _parse_decimal(row.get("torque_20"), "torque_20", row_num, errors, True),
            "torque_40": _parse_decimal(row.get("torque_40"), "torque_40", row_num, errors, False),
            "torque_60": _parse_decimal(row.get("torque_60"), "torque_60", row_num, errors, True),
            "torque_80": _parse_decimal(row.get("torque_80"), "torque_80", row_num, errors, False),
            "torque_100": _parse_decimal(row.get("torque_100"), "torque_100", row_num, errors, True),
            "torque_unit": torque_unit,
            "pressure_20": _parse_decimal(row.get("pressure_20"), "pressure_20", row_num, errors, True),
            "pressure_40": _parse_decimal(row.get("pressure_40"), "pressure_40", row_num, errors, False),
            "pressure_60": _parse_decimal(row.get("pressure_60"), "pressure_60", row_num, errors, True),
            "pressure_80": _parse_decimal(row.get("pressure_80"), "pressure_80", row_num, errors, False),
            "pressure_100": _parse_decimal(row.get("pressure_100"), "pressure_100", row_num, errors, True),
            "pressure_unit": pressure_unit,
            "is_active": _parse_bool(row.get("is_active"), row_num, errors),
        }

        # Enforce range ordering if both values are valid.
        range_min = parsed_data["range_min"]
        range_max = parsed_data["range_max"]
        if range_min is not None and range_max is not None and range_max < range_min:
            errors.append(f"Row {row_num}: range_max must be greater than or equal to range_min")

        objects_to_insert.append(models.HTWManufacturerSpec(**parsed_data))

    if errors:
        raise HTTPException(
            status_code=400,
            detail="Import failed:\n" + "\n".join(errors[:100]),
        )

    try:
        db.add_all(objects_to_insert)
        db.commit()
        return {
            "message": "Import successful",
            "inserted_rows": len(objects_to_insert),
        }
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")

# =====================================================================
# GET: Single HTW Manufacturer Spec (Detail View)
# =====================================================================
@router.get("/{spec_id}", response_model=HTWManufacturerSpecResponse)
def get_htw_manufacturer_spec(
    spec_id: int,
    db: Session = Depends(get_db)
):
    """
    Retrieves a single HTW Manufacturer Spec by ID.
    """
    spec = db.query(models.HTWManufacturerSpec).filter(
        models.HTWManufacturerSpec.id == spec_id
    ).first()

    if not spec:
        raise HTTPException(
            status_code=404,
            detail=f"HTW Manufacturer Spec with ID {spec_id} not found"
        )

    return spec


# =====================================================================
# POST: Create HTW Manufacturer Spec
# =====================================================================
@router.post("/", response_model=HTWManufacturerSpecResponse, status_code=201)
def create_htw_manufacturer_spec(
    spec_data: HTWManufacturerSpecCreate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Creates a new HTW Manufacturer Spec.
    Only works for Hydraulic Torque Wrench equipment type.
    """
    try:
        # Create new spec instance
        new_spec = models.HTWManufacturerSpec(**spec_data.model_dump())

        db.add(new_spec)
        db.commit()
        db.refresh(new_spec)

        return new_spec
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# PUT: Update HTW Manufacturer Spec
# =====================================================================
@router.put("/{spec_id}", response_model=HTWManufacturerSpecResponse)
def update_htw_manufacturer_spec(
    spec_id: int,
    spec_data: HTWManufacturerSpecUpdate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Updates an existing HTW Manufacturer Spec.
    """
    try:
        # Find the spec
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found"
            )

        # Update fields
        update_data = spec_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(spec, field, value)

        # Update the updated_at timestamp
        spec.updated_at = datetime.now()

        db.commit()
        db.refresh(spec)

        return spec
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# PATCH: Update HTW Manufacturer Spec Status
# =====================================================================
@router.patch("/{spec_id}/status", response_model=HTWManufacturerSpecResponse)
def update_htw_manufacturer_spec_status(
    spec_id: int,
    is_active: bool = Query(...),
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Updates the active status of an HTW Manufacturer Spec.
    """
    try:
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found"
            )

        spec.is_active = is_active
        spec.updated_at = datetime.now()

        db.commit()
        db.refresh(spec)

        return spec
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# DELETE: Delete HTW Manufacturer Spec
# =====================================================================
@router.delete("/{spec_id}", status_code=204)
def delete_htw_manufacturer_spec(
    spec_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Deletes an HTW Manufacturer Spec.
    """
    try:
        spec = db.query(models.HTWManufacturerSpec).filter(
            models.HTWManufacturerSpec.id == spec_id
        ).first()

        if not spec:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Manufacturer Spec with ID {spec_id} not found"
            )

        db.delete(spec)
        db.commit()

        return None
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


