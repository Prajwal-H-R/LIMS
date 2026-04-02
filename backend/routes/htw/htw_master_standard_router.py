# backend/routes/htw_master_standard_router.py

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from typing import List, Optional
from datetime import datetime
from io import BytesIO
import pandas as pd
import logging
from starlette.requests import Request
from pydantic import BaseModel

from ...schemas.htw.htw_master_standard_schemas import (
    HTWMasterStandardCreate,
    HTWMasterStandardUpdate,
    HTWMasterStandardResponse
)
from ... import models
from ...db import get_db
from ...auth import get_current_user
from ...schemas.user_schemas import UserResponse


class BatchExportRequest(BaseModel):
    standard_ids: List[int]

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/htw-master-standards",
    tags=["HTW Master Standards"]
)


# =====================================================================
# GET: All HTW Master Standards (List View)
# =====================================================================
@router.get("/", response_model=List[HTWMasterStandardResponse])
def get_htw_master_standards(
    db: Session = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: Optional[int] = Query(None, ge=1),  # no upper bound
    is_active: Optional[bool] = Query(None)
):
    """
    Retrieves a list of HTW Master Standards.
    Only returns standards for Hydraulic Torque Wrench equipment type.
    """
    try:
        query = db.query(models.HTWMasterStandard)
        
        # Filter by active status if provided
        if is_active is not None:
            query = query.filter(models.HTWMasterStandard.is_active == is_active)
        
        # Order by created_at descending (newest first)
        query = query.order_by(models.HTWMasterStandard.created_at.desc())
        
        # Apply pagination
        standards = query.offset(skip).limit(limit).all()
        
        return standards
    except SQLAlchemyError as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# GET: Single HTW Master Standard (Detail View)
# =====================================================================
@router.get("/{standard_id}", response_model=HTWMasterStandardResponse)
def get_htw_master_standard(
    standard_id: int,
    db: Session = Depends(get_db)
):
    """
    Retrieves a single HTW Master Standard by ID.
    """
    standard = db.query(models.HTWMasterStandard).filter(
        models.HTWMasterStandard.id == standard_id
    ).first()
    
    if not standard:
        raise HTTPException(
            status_code=404,
            detail=f"HTW Master Standard with ID {standard_id} not found"
        )
    
    return standard


# =====================================================================
# POST: Create HTW Master Standard
# =====================================================================
@router.post("/", response_model=HTWMasterStandardResponse, status_code=201)
def create_htw_master_standard(
    standard_data: HTWMasterStandardCreate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Creates a new HTW Master Standard.
    Only works for Hydraulic Torque Wrench equipment type.
    """
    try:
        # Create new standard instance
        new_standard = models.HTWMasterStandard(**standard_data.model_dump())
        
        db.add(new_standard)
        db.commit()
        db.refresh(new_standard)
        
        return new_standard
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# PUT: Update HTW Master Standard
# =====================================================================
@router.put("/{standard_id}", response_model=HTWMasterStandardResponse)
def update_htw_master_standard(
    standard_id: int,
    standard_data: HTWMasterStandardUpdate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Updates an existing HTW Master Standard.
    """
    try:
        # Find the standard
        standard = db.query(models.HTWMasterStandard).filter(
            models.HTWMasterStandard.id == standard_id
        ).first()
        
        if not standard:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Master Standard with ID {standard_id} not found"
            )
        
        # Update fields
        update_data = standard_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(standard, field, value)
        
        # Update the updated_at timestamp
        standard.updated_at = datetime.now()
        
        db.commit()
        db.refresh(standard)
        
        return standard
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# PATCH: Update HTW Master Standard Status
# =====================================================================
@router.patch("/{standard_id}/status", response_model=HTWMasterStandardResponse)
def update_htw_master_standard_status(
    standard_id: int,
    is_active: bool = Query(...),
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Updates the active status of an HTW Master Standard.
    """
    try:
        standard = db.query(models.HTWMasterStandard).filter(
            models.HTWMasterStandard.id == standard_id
        ).first()
        
        if not standard:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Master Standard with ID {standard_id} not found"
            )
        
        standard.is_active = is_active
        standard.updated_at = datetime.now()
        
        db.commit()
        db.refresh(standard)
        
        return standard
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# DELETE: Delete HTW Master Standard
# =====================================================================
@router.delete("/{standard_id}", status_code=204)
def delete_htw_master_standard(
    standard_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Deletes an HTW Master Standard.
    """
    try:
        standard = db.query(models.HTWMasterStandard).filter(
            models.HTWMasterStandard.id == standard_id
        ).first()
        
        if not standard:
            raise HTTPException(
                status_code=404,
                detail=f"HTW Master Standard with ID {standard_id} not found"
            )
        
        db.delete(standard)
        db.commit()
        
        return None
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An internal server error occurred: {str(e)}")


# =====================================================================
# GET: Export HTW Master Standards to Excel
# =====================================================================
@router.get("/export")
def export_htw_master_standards(
    request: Request,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Exports HTW Master Standards to Excel format.
    Can export all standards or a specific list of standards by IDs.
    Accepts multiple 'ids' query parameters: ?ids=1&ids=2&ids=3
    Accepts optional 'is_active' query parameter: ?is_active=true
    Similar to the inward export feature in engineering portal.
    """
    try:
        # Manually parse all query parameters to avoid FastAPI validation issues
        ids_list = request.query_params.getlist("ids")
        ids = None
        if ids_list:
            try:
                ids = [int(id_str) for id_str in ids_list if id_str and id_str.strip()]
                if ids:
                    logger.info(f"Exporting {len(ids)} selected master standards: {ids}")
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid ID format in export request: {ids_list}, error: {e}")
                raise HTTPException(status_code=400, detail="Invalid ID format. IDs must be integers.")
        
        # Parse is_active parameter manually
        is_active_param = request.query_params.get("is_active")
        is_active = None
        if is_active_param is not None:
            if is_active_param.lower() in ("true", "1", "yes"):
                is_active = True
            elif is_active_param.lower() in ("false", "0", "no"):
                is_active = False
        
        # Query standards
        query = db.query(models.HTWMasterStandard)
        
        # Filter by IDs if provided
        if ids and len(ids) > 0:
            query = query.filter(models.HTWMasterStandard.id.in_(ids))
        else:
            logger.info("Exporting all master standards (no IDs specified)")
        
        # Filter by active status if provided
        if is_active is not None:
            query = query.filter(models.HTWMasterStandard.is_active == is_active)
        
        standards = query.order_by(models.HTWMasterStandard.created_at.desc()).all()
        
        if not standards:
            logger.warning("No master standards found for export")
            raise HTTPException(status_code=404, detail="No master standards found to export.")
        
        logger.info(f"Exporting {len(standards)} master standard(s)")
        
        # Prepare data for Excel - Include ALL columns from the table
        rows = []
        for std in standards:
            # Calculate calibration status based on calibration_valid_upto
            calibration_status = "N/A"
            if std.calibration_valid_upto:
                today = datetime.now().date()
                if std.calibration_valid_upto < today:
                    calibration_status = "Expired"
                else:
                    calibration_status = "Active"
            
            # Build row with ALL columns from htw_master_standard table
            row = {
                "ID": std.id,
                "Nomenclature": std.nomenclature or "",
                "Range Min": float(std.range_min) if std.range_min is not None else "",
                "Range Max": float(std.range_max) if std.range_max is not None else "",
                "Range Unit": std.range_unit or "",
                "Manufacturer": std.manufacturer or "",
                "Model / Serial No": std.model_serial_no or "",
                "Traceable To Lab": std.traceable_to_lab or "",
                "Uncertainty": float(std.uncertainty) if std.uncertainty is not None else "",
                "Uncertainty Unit": std.uncertainty_unit or "",
                "Certificate No": std.certificate_no or "",
                "Calibration Valid Upto": std.calibration_valid_upto.strftime("%Y-%m-%d") if std.calibration_valid_upto else "",
                "Calibration Status": calibration_status,  # Calculated field
                "Accuracy of Master": std.accuracy_of_master or "",
                "Resolution": float(std.resolution) if std.resolution is not None else "",
                "Resolution Unit": std.resolution_unit or "",
                "Is Active": "Yes" if std.is_active else "No",
                "Created At": std.created_at.strftime("%Y-%m-%d %H:%M:%S") if std.created_at else "",
                "Updated At": std.updated_at.strftime("%Y-%m-%d %H:%M:%S") if std.updated_at else "",
            }
            rows.append(row)
        
        # Create DataFrame with pandas
        df = pd.DataFrame(rows)
        
        # Generate Excel file using pandas and openpyxl (similar to inward export)
        output = BytesIO()
        try:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Write DataFrame to Excel
                df.to_excel(writer, index=False, sheet_name="Master Standards")
                
                # Get the workbook and worksheet for formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                worksheet = writer.sheets["Master Standards"]
                
                # Auto-adjust column widths (similar to inward export pattern)
                for idx, col in enumerate(df.columns, 1):
                    max_length = max(
                        df[col].astype(str).map(len).max(),  # Max length in data
                        len(str(col))  # Length of column name
                    )
                    # Set column width (add some padding, max 50)
                    column_letter = get_column_letter(idx)
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Format header row (bold, background color - matching inward export style)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Freeze header row
                worksheet.freeze_panes = "A2"
            
            output.seek(0)
            
            # Generate filename with timestamp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"htw_master_standards_export_{timestamp}.xlsx"
            
            logger.info(f"Successfully generated export file: {filename}")
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        except Exception as excel_error:
            logger.error(f"Failed to generate Excel file: {excel_error}", exc_info=True)
            raise HTTPException(status_code=500, detail="Failed to generate Excel export file.")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export master standards: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate export: {str(e)}")


# =====================================================================
# POST: Batch Export HTW Master Standards to Excel
# =====================================================================
@router.post("/export-batch")
def export_htw_master_standards_batch(
    request_data: BatchExportRequest,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Exports multiple HTW Master Standards to Excel format.
    Accepts a list of standard IDs in the request body.
    Similar to the inward batch export feature.
    """
    try:
        if not request_data.standard_ids or len(request_data.standard_ids) == 0:
            raise HTTPException(status_code=400, detail="No standard IDs provided for export.")
        
        logger.info(f"Batch exporting {len(request_data.standard_ids)} master standards: {request_data.standard_ids}")
        
        # Query standards by IDs
        standards = db.query(models.HTWMasterStandard).filter(
            models.HTWMasterStandard.id.in_(request_data.standard_ids)
        ).order_by(models.HTWMasterStandard.created_at.desc()).all()
        
        if not standards:
            logger.warning(f"No master standards found for IDs: {request_data.standard_ids}")
            raise HTTPException(status_code=404, detail="No master standards found to export.")
        
        logger.info(f"Exporting {len(standards)} master standard(s)")
        
        # Prepare data for Excel - Include ALL columns from the table
        rows = []
        for std in standards:
            # Calculate calibration status based on calibration_valid_upto
            calibration_status = "N/A"
            if std.calibration_valid_upto:
                today = datetime.now().date()
                if std.calibration_valid_upto < today:
                    calibration_status = "Expired"
                else:
                    calibration_status = "Active"
            
            # Build row with ALL columns from htw_master_standard table
            row = {
                "ID": std.id,
                "Nomenclature": std.nomenclature or "",
                "Range Min": float(std.range_min) if std.range_min is not None else "",
                "Range Max": float(std.range_max) if std.range_max is not None else "",
                "Range Unit": std.range_unit or "",
                "Manufacturer": std.manufacturer or "",
                "Model / Serial No": std.model_serial_no or "",
                "Traceable To Lab": std.traceable_to_lab or "",
                "Uncertainty": float(std.uncertainty) if std.uncertainty is not None else "",
                "Uncertainty Unit": std.uncertainty_unit or "",
                "Certificate No": std.certificate_no or "",
                "Calibration Valid Upto": std.calibration_valid_upto.strftime("%Y-%m-%d") if std.calibration_valid_upto else "",
                "Calibration Status": calibration_status,  # Calculated field
                "Accuracy of Master": std.accuracy_of_master or "",
                "Resolution": float(std.resolution) if std.resolution is not None else "",
                "Resolution Unit": std.resolution_unit or "",
                "Is Active": "Yes" if std.is_active else "No",
                "Created At": std.created_at.strftime("%Y-%m-%d %H:%M:%S") if std.created_at else "",
                "Updated At": std.updated_at.strftime("%Y-%m-%d %H:%M:%S") if std.updated_at else "",
            }
            rows.append(row)
        
        # Create DataFrame with pandas
        df = pd.DataFrame(rows)
        
        # Generate Excel file using pandas and openpyxl (similar to GET export)
        output = BytesIO()
        try:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Write DataFrame to Excel
                df.to_excel(writer, index=False, sheet_name="Master Standards")
                
                # Get the workbook and worksheet for formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                worksheet = writer.sheets["Master Standards"]
                
                # Auto-adjust column widths
                for idx, col in enumerate(df.columns, 1):
                    max_length = max(
                        df[col].astype(str).map(len).max(),  # Max length in data
                        len(str(col))  # Length of column name
                    )
                    # Set column width (add some padding, max 50)
                    column_letter = get_column_letter(idx)
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Format header row (bold, background color)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Freeze header row
                worksheet.freeze_panes = "A2"
            
            output.seek(0)
            
            # Generate filename with timestamp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"htw_master_standards_export_{timestamp}.xlsx"
            
            logger.info(f"Successfully generated batch export file: {filename}")
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        except Exception as excel_error:
            logger.error(f"Failed to generate Excel file: {excel_error}", exc_info=True)
            raise HTTPException(status_code=500, detail="Failed to generate Excel export file.")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to batch export master standards: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate batch export: {str(e)}")

